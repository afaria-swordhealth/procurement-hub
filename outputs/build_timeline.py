import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SWORD_FILL  = PatternFill("solid", fgColor="2E75B6")
TRANS_FILL  = PatternFill("solid", fgColor="70AD47")
BOTH_FILL   = PatternFill("solid", fgColor="ED7D31")
NOTE_FILL   = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL    = PatternFill("solid", fgColor="D6E4F7")
EDIT_FILL   = PatternFill("solid", fgColor="FFFF99")

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def build_sheet(ws, title, steps, footnote):
    ws.title = title
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 48

    # Title row
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = title
    t.font = Font(bold=True, size=14, color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    headers = ["Step", "Task", "Owner", "Start Date", "Duration (days)", "End Date", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 22

    for i, (step, task, owner, start_val, dur, notes) in enumerate(steps):
        r = 3 + i
        ws.row_dimensions[r].height = 22
        row_fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        owner_fill = SWORD_FILL if owner == "Sword" else (TRANS_FILL if owner == "Transtek" else BOTH_FILL)
        owner_font = Font(color="FFFFFF", size=10, bold=True)

        def wc(col, val=None, formula=None, num_fmt=None, fill=None, fnt=None, align=None):
            c = ws.cell(row=r, column=col)
            c.value = formula if formula else val
            if num_fmt:
                c.number_format = num_fmt
            c.fill = fill or row_fill
            c.font = fnt or Font(size=10)
            c.alignment = align or Alignment(vertical="center", wrap_text=(col == 7))
            c.border = border

        wc(1, step, align=Alignment(horizontal="center", vertical="center"))
        wc(2, task)
        # Owner
        oc = ws.cell(row=r, column=3, value=owner)
        oc.fill = owner_fill
        oc.font = owner_font
        oc.alignment = Alignment(horizontal="center", vertical="center")
        oc.border = border
        # Start date
        sc = ws.cell(row=r, column=4)
        if isinstance(start_val, str) and start_val.startswith("="):
            sc.value = start_val
        else:
            sc.value = start_val
        sc.number_format = "DD-MMM-YYYY"
        sc.font = Font(size=10)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.fill = row_fill
        sc.border = border
        # Duration (editable)
        dc = ws.cell(row=r, column=5, value=dur)
        dc.fill = EDIT_FILL
        dc.font = Font(size=10, bold=True)
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = border
        # End date
        ec = ws.cell(row=r, column=6, value=f"=D{r}+E{r}-1")
        ec.number_format = "DD-MMM-YYYY"
        ec.font = Font(size=10)
        ec.alignment = Alignment(horizontal="center", vertical="center")
        ec.fill = row_fill
        ec.border = border
        # Notes
        wc(7, notes, fill=NOTE_FILL, align=Alignment(vertical="center", wrap_text=True))

    # Footnote
    fn_row = 3 + len(steps) + 1
    ws.merge_cells(f"A{fn_row}:G{fn_row}")
    fn = ws[f"A{fn_row}"]
    fn.value = footnote
    fn.font = Font(italic=True, size=9, color="666666")
    fn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[fn_row].height = 30

    # Legend
    lr = fn_row + 2
    ws.cell(row=lr, column=1, value="Legend:").font = Font(bold=True, size=9)
    for col, (label, fill) in enumerate([("Sword", SWORD_FILL), ("Transtek", TRANS_FILL), ("Both", BOTH_FILL), ("Editable", EDIT_FILL)], 2):
        lc = ws.cell(row=lr, column=col, value=label)
        lc.fill = fill
        lc.font = Font(color="FFFFFF" if label != "Editable" else "000000", size=9, bold=True)
        lc.alignment = Alignment(horizontal="center")
        lc.border = border


apr24 = date(2026, 4, 24)

# ── OTS BPM ──────────────────────────────────────────────────────────────────
bpm_steps = [
    ("1",  "SQA review, discussion & approval signature",
     "Both",  apr24, 5,
     "Sword feedback by Apr 27. Transtek to add Scale UDI-DI by Apr 29."),
    ("2",  "Confirm order details & deposit payment (50%)",
     "Sword", apr24, 13,
     "PO + invoice + 50% deposit to be settled. Runs in parallel with Step 1."),
    ("3",  "Mass production",
     "Transtek", "=MAX(F3,F4)+1", 28,
     "Starts after SQA approval AND deposit cleared. Jun 19-21 = CN holiday (absorbed in duration)."),
    ("4a", "Air shipment — China to Salt Lake City",
     "Sword", "=F5+1", 15,
     "Arranged by Sword shipping agent. ~15 days estimated."),
    ("4b", "Sea shipment — China to Salt Lake City (alt)",
     "Sword", "=F5+1", 45,
     "Parallel alternative. ~45 days. Lower cost. Adjust duration in E to model scenarios."),
]
bpm_note = (
    "Yellow cells = editable duration. Changing a duration auto-updates all downstream dates.  "
    "Phases removed: Packaging sample production | Golden sample | Golden sample DHL approval | Quality inspection."
)
ws_bpm = wb.active
build_sheet(ws_bpm, "OTS BPM — BB2284-AE01", bpm_steps, bpm_note)

# ── OTS Scale ─────────────────────────────────────────────────────────────────
scale_steps = [
    ("1",  "SQA review, discussion & approval signature",
     "Both",  apr24, 5,
     "Scale UDI-DI confirmation from Transtek required by Apr 29. Regulatory decision on scale track."),
    ("2",  "Confirm scale order details & deposit payment (50%)",
     "Sword", apr24, 13,
     "PO + invoice + 50% deposit. 2,000 units GBF-2008-B1. Runs in parallel with Step 1."),
    ("3",  "Scale mass production",
     "Transtek", "=MAX(F3,F4)+1", 28,
     "Starts after SQA approval AND deposit cleared. Jun 19-21 = CN holiday (absorbed). Note: no OTS scale stock available."),
    ("4a", "Air shipment — China to Salt Lake City",
     "Sword", "=F5+1", 15,
     "Arranged by Sword shipping agent. ~15 days estimated."),
    ("4b", "Sea shipment — China to Salt Lake City (alt)",
     "Sword", "=F5+1", 45,
     "Parallel alternative. ~45 days. Lower cost. Adjust duration in E to model scenarios."),
]
scale_note = (
    "Yellow cells = editable duration. Changing a duration auto-updates all downstream dates.  "
    "Note: 0 OTS units in stock for scale (GBF-2008-B1). This plan reflects the fast-track branded path without golden sample phases."
)
ws_scale = wb.create_sheet("OTS Scale — GBF-2008-B1")
build_sheet(ws_scale, "OTS Scale — GBF-2008-B1", scale_steps, scale_note)

out = r"G:\My Drive\Projects\Pulse\Device Procurement\BPM\Pulse_OTS_Timeline_Dynamic_20260424.xlsx"
wb.save(out)
print("Saved:", out)
