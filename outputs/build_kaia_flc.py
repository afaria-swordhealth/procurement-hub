"""
Kaia FLC updater — Apr 25, 2026
Updates Second Page Fernando confirmed quotes + adds 3mm Scenarios sheet.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = "G:/My Drive/Projects/Kaia Rewards/Kaia_FLC.xlsx"

# ── colour palette ────────────────────────────────────────────────────
BLUE_H   = PatternFill("solid", fgColor="1F4E79")   # dark blue header
BLUE_L   = PatternFill("solid", fgColor="BDD7EE")   # light blue section
ORANGE_H = PatternFill("solid", fgColor="F4B942")   # Tiger header
ORANGE_L = PatternFill("solid", fgColor="FCE4D6")   # Tiger rows
GREEN_H  = PatternFill("solid", fgColor="375623")   # Second Page header
GREEN_L  = PatternFill("solid", fgColor="E2EFDA")   # Second Page rows
GREY_L   = PatternFill("solid", fgColor="F2F2F2")   # label rows
YELLOW   = PatternFill("solid", fgColor="FFFF99")   # FLC results
WHITE    = PatternFill("solid", fgColor="FFFFFF")

BOLD_W  = Font(bold=True, color="FFFFFF")
BOLD_B  = Font(bold=True, color="000000")
NORM    = Font(color="000000")
ITALIC  = Font(italic=True, color="595959", size=9)

def hdr(fill, font=None):
    return {"fill": fill, "font": font or BOLD_W, "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)}

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, fill=None, font=None, align=None, number_format=None, border=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:    c.fill = fill
    if font:    c.font = font
    if align:   c.alignment = align
    if number_format: c.number_format = number_format
    if border:  c.border = thin_border()
    return c

USD = '"$"#,##0.00'
USD3 = '"$"#,##0.000'

# ── Nimbl constants ───────────────────────────────────────────────────
NIMBL_LOW  = 13.15
NIMBL_HIGH = 17.15
BASELINE   = 27.00   # 4imprint + SV Direct low

# ── Tiger Fitness 3mm freight + duty data ─────────────────────────────
TF = {
    # unit prices
    "price_3k": 3.20, "price_5k": 2.80,
    # carton: 12 pcs, 38×30×25 cm, 10 kg
    "ctns_3k": 250, "cbm_3k": 7.20, "kg_3k": 2500,
    "ctns_5k": 417, "cbm_5k": 11.90, "kg_5k": 4167,
    # Fernando confirmed (Slack Mar 12 / Mar 17)
    "sea_lcl_3k": 4900, "sea_lcl_5k": 5950,
    "air_3k":    12700, "air_5k":    20000,
    # Eva supplier CNF to US port (Apr 8, ambiguous — A used)
    "eva_cnf_5k_a": 4998, "eva_cnf_5k_b": 6555,
    # FCL: not quoted by Fernando for Tiger
}

# ── Second Page 3mm freight + duty data ───────────────────────────────
SP = {
    "price_3k": 2.59, "price_5k": 2.39,
    # carton: 20 pcs, 46×36×63 cm, ~16 kg
    "ctns_3k": 150, "cbm_3k": 15.65, "kg_3k": 2400,
    "ctns_5k": 250, "cbm_5k": 26.00, "kg_5k": 4000,
    # Fernando CONFIRMED (Slack Apr 13-15 — new, not previously in file)
    "sea_lcl_3k": 5990, "sea_lcl_5k": 8102,
    "sea_fcl_5k": 7966,              # 20' FCL door/door
    "air_3k":    19487, "air_5k":    29783,
    # Jerry DDP to SLC (indicative Mar 16 — incl. duties)
    "jerry_ddp_per_unit": 1.25,
}

# ── duty calc helper ──────────────────────────────────────────────────
def duties(unit_price, qty, mode="sea"):
    merch  = unit_price * qty
    mfn    = merch * 0.046
    s301   = merch * 0.100
    mpf    = max(31.67, min(614.35, merch * 0.003464))
    hmf    = merch * 0.00125 if mode == "sea" else 0.0
    total  = mfn + s301 + mpf + hmf
    return total, total / qty

def flc(unit, freight_per_unit, duty_per_unit):
    landed = unit + freight_per_unit + duty_per_unit
    return landed, landed + NIMBL_LOW, landed + NIMBL_HIGH

# ── pre-compute all scenarios ─────────────────────────────────────────
# Tiger
tf_sea_3k  = duties(TF["price_3k"], 3000, "sea")
tf_sea_5k  = duties(TF["price_5k"], 5000, "sea")
tf_air_3k  = duties(TF["price_3k"], 3000, "air")
tf_air_5k  = duties(TF["price_5k"], 5000, "air")

# Second Page
sp_sea_3k  = duties(SP["price_3k"], 3000, "sea")
sp_sea_5k  = duties(SP["price_5k"], 5000, "sea")
sp_air_3k  = duties(SP["price_3k"], 3000, "air")
sp_air_5k  = duties(SP["price_5k"], 5000, "air")

# ─────────────────────────────────────────────────────────────────────
# Scenarios dict: (freight_total, freight_per_unit, duty_total, duty_per_unit)
# DDP rows: duties = 0 (already included by supplier)
# ─────────────────────────────────────────────────────────────────────
def scene(frght_total, unit_price, qty, mode="sea"):
    fpu = frght_total / qty
    dt, dpu = duties(unit_price, qty, mode)
    landed, flc_lo, flc_hi = flc(unit_price, fpu, dpu)
    return frght_total, fpu, dt, dpu, landed, flc_lo, flc_hi

def scene_ddp(ddp_per_unit, unit_price, qty):
    """Supplier DDP: duties included in DDP price, no separate duty line."""
    landed  = unit_price + ddp_per_unit
    flc_lo  = landed + NIMBL_LOW
    flc_hi  = landed + NIMBL_HIGH
    return ddp_per_unit * qty, ddp_per_unit, 0, 0, landed, flc_lo, flc_hi

# Tiger rows
TF_ROWS = {
    "@3K Fernando Sea LCL":  scene(TF["sea_lcl_3k"], TF["price_3k"], 3000, "sea"),
    "@3K Fernando Air":      scene(TF["air_3k"],     TF["price_3k"], 3000, "air"),
    "@5K Eva CNF (to port)*":scene(TF["eva_cnf_5k_a"], TF["price_5k"], 5000, "sea"),
    "@5K Fernando Sea LCL":  scene(TF["sea_lcl_5k"], TF["price_5k"], 5000, "sea"),
    "@5K Fernando Air":      scene(TF["air_5k"],     TF["price_5k"], 5000, "air"),
}
# Second Page rows
SP_ROWS = {
    "@3K Jerry DDP SLC†":    scene_ddp(SP["jerry_ddp_per_unit"], SP["price_3k"], 3000),
    "@3K Fernando Sea LCL":  scene(SP["sea_lcl_3k"], SP["price_3k"], 3000, "sea"),
    "@3K Fernando Air":      scene(SP["air_3k"],     SP["price_3k"], 3000, "air"),
    "@5K Jerry DDP SLC†":    scene_ddp(SP["jerry_ddp_per_unit"], SP["price_5k"], 5000),
    "@5K Fernando Sea LCL":  scene(SP["sea_lcl_5k"], SP["price_5k"], 5000, "sea"),
    "@5K Fernando Sea FCL":  scene(SP["sea_fcl_5k"], SP["price_5k"], 5000, "sea"),
    "@5K Fernando Air":      scene(SP["air_5k"],     SP["price_5k"], 5000, "air"),
}

# ══════════════════════════════════════════════════════════════════════
wb = openpyxl.load_workbook(PATH)

# ── 1. Update Second Page Yoga sheet (add Fernando confirmed section) ─
ws_sp = wb["Second Page Yoga"]

# Find first empty row after current content
start_row = ws_sp.max_row + 2

def write_section_sp(ws, r):
    # Header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(r, 1, "4b. FREIGHT — FERNANDO SARAIVA CONFIRMED (Slack Apr 13-15, 2026)")
    c.fill = GREEN_H; c.font = BOLD_W
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(r, 1, "Door/Door China → SLC (Nimbl). Excludes import duties. Supersedes Jerry DDP as independent benchmark.")
    c.fill = GREEN_L; c.font = ITALIC
    r += 1

    heads = ["", "@3K", "@5K", "Per unit @3K", "Per unit @5K", "Notes"]
    widths = [30, 12, 12, 14, 14, 45]
    for i, h in enumerate(heads, 1):
        c = ws.cell(r, i, h); c.fill = BLUE_L; c.font = BOLD_B
        c.alignment = Alignment(horizontal="center")
    r += 1

    rows = [
        ("Sea LCL (door/door, ex duties)", SP["sea_lcl_3k"], SP["sea_lcl_5k"],
         SP["sea_lcl_3k"]/3000, SP["sea_lcl_5k"]/5000,
         "LCL = Less-than-container load. 25-35d transit. Fernando, Apr 13."),
        ("Sea FCL 20' (door/door, ex duties)", None, SP["sea_fcl_5k"],
         None, SP["sea_fcl_5k"]/5000,
         "@5K only. Full container (26 cbm fills 20'). Marginally cheaper than LCL. Fernando, Apr 15."),
        ("Air freight (door/door, ex duties)", SP["air_3k"], SP["air_5k"],
         SP["air_3k"]/3000, SP["air_5k"]/5000,
         "5-8d transit. Use for urgent only — 3x cost of sea. Fernando, Apr 13."),
    ]
    for label, t3k, t5k, p3k, p5k, note in rows:
        ws.cell(r, 1, label).font = NORM
        if t3k: ws.cell(r, 2, t3k).number_format = USD
        if t5k: ws.cell(r, 3, t5k).number_format = USD
        if p3k: ws.cell(r, 4, p3k).number_format = USD3
        if p5k: ws.cell(r, 5, p5k).number_format = USD3
        ws.cell(r, 6, note).font = ITALIC
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(r, 1,
        "Note: Fernando's LCL sea @5K = $1.62/unit vs Jerry DDP $1.25/unit. "
        "Fernando +duties ($0.35) = $1.97 total vs Jerry DDP $1.25 all-in. "
        "Jerry's DDP was flagged as low by Fernando (Mar 17). Confirm with Jerry before PO.")
    c.font = ITALIC; c.fill = GREY_L
    return r + 2

write_section_sp(ws_sp, start_row)

# ── 2. Create "3mm Scenarios" sheet ───────────────────────────────────
if "3mm Scenarios" in wb.sheetnames:
    del wb["3mm Scenarios"]
ws3 = wb.create_sheet("3mm Scenarios")

# column widths
col_widths = [32, 14, 14, 14, 14, 14, 14, 14, 14, 14]
for i, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = 1

# ── Title ─────────────────────────────────────────────────────────────
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws3.cell(r, 1, "Kaia Yoga Mat — 3mm Full Landed Cost Scenarios")
c.fill = BLUE_H; c.font = Font(bold=True, size=14, color="FFFFFF")
c.alignment = Alignment(horizontal="center")
r += 1

ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws3.cell(r, 1,
    "Freight options: Supplier-arranged  |  Fernando Saraiva (ISC Logistics) — Sea LCL  |  Fernando — Sea FCL  |  Fernando — Air  "
    " |  Updated: Apr 25, 2026")
c.font = ITALIC; c.fill = GREY_L
c.alignment = Alignment(horizontal="center")
r += 1

ws3.row_dimensions[r].height = 8
r += 1

# ── Column headers ────────────────────────────────────────────────────
col_heads = [
    "Scenario",
    "Freight\nTotal ($)",
    "Freight\n/unit ($)",
    "Duties\nTotal ($)",
    "Duties\n/unit ($)",
    "Landed\n(ex-Nimbl)",
    "FLC Low\n(+Nimbl $13.15)",
    "FLC High\n(+Nimbl $17.15)",
    "vs Baseline\n($27 low)",
    "Notes"
]
for i, h in enumerate(col_heads, 1):
    c = ws3.cell(r, i, h)
    c.fill = BLUE_H; c.font = BOLD_W
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[r].height = 36
r += 1

# ── Helper to write a supplier block ──────────────────────────────────
def write_supplier_block(ws, r, name, unit_prices, rows_data, hdr_fill, row_fill):
    # Supplier header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(r, 1, name)
    c.fill = hdr_fill; c.font = BOLD_W
    r += 1

    for label, unit_price, (ft, fpu, dt, dpu, landed, flc_lo, flc_hi), note in rows_data:
        cells = [label, ft or "", fpu or "", dt or "", dpu or "",
                 landed, flc_lo, flc_hi, BASELINE - flc_lo, note]
        for i, val in enumerate(cells, 1):
            c = ws.cell(r, i, val if val != "" else None)
            c.fill = row_fill
            if i in (1,): c.font = BOLD_B
            if i == 1:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif i == 10:
                c.font = ITALIC; c.alignment = Alignment(wrap_text=True)
            elif i in (6, 7, 8, 9):
                c.fill = YELLOW if i in (7, 8) else row_fill
                c.font = Font(bold=True) if i in (7, 8) else NORM
                if val and isinstance(val, (int, float)):
                    c.number_format = USD
            elif val and isinstance(val, (int, float)):
                c.number_format = USD
        r += 1
    return r

# ── Tiger Fitness rows ────────────────────────────────────────────────
tf_note_sea = "Fernando confirmed (Slack Mar 17). Door/door ex duties. 25-35d transit."
tf_note_air = "Fernando confirmed (Slack Mar 12). Door/door ex duties. 5-8d transit."
tf_note_eva = "Eva (Tiger) CNF to US port, customer pickup. +last-mile to SLC not included. Ambiguous: may be $6,555 alt. Verify."

r = write_supplier_block(ws3, r,
    "TIGER FITNESS  (Shanghai, CN)  |  EXW Shanghai  |  HTS 9506.91.0030  |  Duties: 4.6% MFN + 10% Section 301",
    None,
    [
        ("@3K — Fernando Sea LCL",  TF["price_3k"], scene(TF["sea_lcl_3k"], TF["price_3k"], 3000, "sea"), tf_note_sea),
        ("@3K — Fernando Air",      TF["price_3k"], scene(TF["air_3k"],     TF["price_3k"], 3000, "air"), tf_note_air),
        ("@5K — Eva CNF to port *", TF["price_5k"], scene(TF["eva_cnf_5k_a"], TF["price_5k"], 5000, "sea"), tf_note_eva),
        ("@5K — Fernando Sea LCL",  TF["price_5k"], scene(TF["sea_lcl_5k"], TF["price_5k"], 5000, "sea"), tf_note_sea),
        ("@5K — Fernando Air",      TF["price_5k"], scene(TF["air_5k"],     TF["price_5k"], 5000, "air"), tf_note_air),
    ],
    ORANGE_H, ORANGE_L
)

ws3.row_dimensions[r].height = 8
r += 1

# ── Second Page Yoga rows ─────────────────────────────────────────────
sp_note_ddp  = "Jerry Ding (Second Page) indicative DDP SLC, incl. duties. ~35d. Flagged as low by Fernando — confirm before PO."
sp_note_sea  = "Fernando confirmed (Slack Apr 13-15). Door/door ex duties. 25-35d transit."
sp_note_fcl  = "Fernando confirmed (Slack Apr 15). 20' FCL door/door ex duties. @5K only — 26 cbm fills 20' container."
sp_note_air  = "Fernando confirmed (Slack Apr 13). Door/door ex duties. 5-8d transit."

r = write_supplier_block(ws3, r,
    "SECOND PAGE YOGA  (Hefei, CN)  |  EXW Hefei / Port Ningbo  |  HTS 9506.91.0030  |  Duties: 4.6% MFN + 10% Section 301",
    None,
    [
        ("@3K — Jerry DDP SLC †",   SP["price_3k"], scene_ddp(SP["jerry_ddp_per_unit"], SP["price_3k"], 3000), sp_note_ddp),
        ("@3K — Fernando Sea LCL",  SP["price_3k"], scene(SP["sea_lcl_3k"], SP["price_3k"], 3000, "sea"), sp_note_sea),
        ("@3K — Fernando Air",      SP["price_3k"], scene(SP["air_3k"],     SP["price_3k"], 3000, "air"), sp_note_air),
        ("@5K — Jerry DDP SLC †",   SP["price_5k"], scene_ddp(SP["jerry_ddp_per_unit"], SP["price_5k"], 5000), sp_note_ddp),
        ("@5K — Fernando Sea LCL",  SP["price_5k"], scene(SP["sea_lcl_5k"], SP["price_5k"], 5000, "sea"), sp_note_sea),
        ("@5K — Fernando Sea FCL",  SP["price_5k"], scene(SP["sea_fcl_5k"], SP["price_5k"], 5000, "sea"), sp_note_fcl),
        ("@5K — Fernando Air",      SP["price_5k"], scene(SP["air_5k"],     SP["price_5k"], 5000, "air"), sp_note_air),
    ],
    GREEN_H, GREEN_L
)

ws3.row_dimensions[r].height = 8
r += 1

# ── Assumptions + Legend ──────────────────────────────────────────────
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c = ws3.cell(r, 1, "ASSUMPTIONS & NOTES")
c.fill = BLUE_H; c.font = BOLD_W
r += 1

assumptions = [
    "Nimbl fulfillment (low $13.15 / high $17.15): ISC negotiated rates. Includes order processing, handling, medium box, outbound shipping.",
    "Baseline: 4imprint + SV Direct = $27-$31/unit. Savings vs baseline = $27 - FLC Low.",
    "Duties: HTS 9506.91.0030 (yoga/exercise mat). MFN 4.6% + Section 301 (China List 4A) 10% + MPF (min $31.67, max $614.35) + HMF (sea only, 0.125%).",
    "* Tiger Eva CNF: supplier-quoted freight to US port (approx. LAX). Customer pickup — excludes last-mile to Nimbl SLC. Add ~$200-400/shipment inland freight.",
    "† Second Page Jerry DDP: indicative rate, includes all freight + duties to SLC. Fernando flagged as potentially low (Mar 17). Confirm with Jerry before using as FLC basis.",
    "FCL (Full Container Load) only available for 5K+ when 20-25 cbm fills a 20' container. 3K Second Page = 15.65 cbm — borderline; can request FCL quote.",
    "Fernando quotes: door-to-door, exclude duties. All quotes confirmed via Slack. Tiger: Mar 12-17. Second Page: Apr 13-15.",
]
for a in assumptions:
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws3.cell(r, 1, a)
    c.fill = GREY_L; c.font = ITALIC
    r += 1

# ── Move sheet to position 2 (after Summary) ─────────────────────────
wb.move_sheet("3mm Scenarios", offset=-(len(wb.sheetnames)-2))

wb.save(PATH)
print("DONE — Kaia_FLC.xlsx updated.")
print(f"Sheets: {wb.sheetnames}")

# ── Print scenario summary to console ────────────────────────────────
print("\n=== 3mm FLC SUMMARY ===")
print(f"{'Scenario':<38} {'FLC Low':>9} {'FLC High':>9} {'vs $27':>7}")
print("-"*70)
for label, (ft, fpu, dt, dpu, landed, flc_lo, flc_hi) in [
    ("Tiger @3K Fernando Sea LCL",  scene(TF["sea_lcl_3k"], TF["price_3k"], 3000, "sea")),
    ("Tiger @3K Fernando Air",      scene(TF["air_3k"],     TF["price_3k"], 3000, "air")),
    ("Tiger @5K Eva CNF+port",      scene(TF["eva_cnf_5k_a"], TF["price_5k"], 5000, "sea")),
    ("Tiger @5K Fernando Sea LCL",  scene(TF["sea_lcl_5k"], TF["price_5k"], 5000, "sea")),
    ("Tiger @5K Fernando Air",      scene(TF["air_5k"],     TF["price_5k"], 5000, "air")),
    ("SP @3K Jerry DDP",            scene_ddp(SP["jerry_ddp_per_unit"], SP["price_3k"], 3000)),
    ("SP @3K Fernando Sea LCL",     scene(SP["sea_lcl_3k"], SP["price_3k"], 3000, "sea")),
    ("SP @3K Fernando Air",         scene(SP["air_3k"],     SP["price_3k"], 3000, "air")),
    ("SP @5K Jerry DDP",            scene_ddp(SP["jerry_ddp_per_unit"], SP["price_5k"], 5000)),
    ("SP @5K Fernando Sea LCL",     scene(SP["sea_lcl_5k"], SP["price_5k"], 5000, "sea")),
    ("SP @5K Fernando Sea FCL",     scene(SP["sea_fcl_5k"], SP["price_5k"], 5000, "sea")),
    ("SP @5K Fernando Air",         scene(SP["air_5k"],     SP["price_5k"], 5000, "air")),
]:
    print(f"  {label:<36} ${flc_lo:>7.2f}   ${flc_hi:>7.2f}   ${BASELINE-flc_lo:>5.2f}")
