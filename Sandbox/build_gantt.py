import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pulse PLD Gantt"

# Colors
colors = {
    "header": "1F222C",
    "prepo": "7BA7D1",
    "design": "E8A5A0",
    "approval": "E5A34B",
    "prod_bpm": "5B8DB8",
    "prod_scale": "4CAF7D",
    "fda": "A4AAB6",
    "finance": "D35D5D",
}

border_thin = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)

# Timeline: weeks from Apr 7 to Nov 2 (30 weeks)
start_date = datetime(2026, 4, 6)
num_weeks = 30
weeks = [start_date + timedelta(weeks=w) for w in range(num_weeks)]

# Task info columns
info_cols = ["#", "Track", "Task", "Owner", "Duration", "Depends On"]
info_widths = [4, 10, 38, 18, 12, 16]
gantt_start_col = len(info_cols) + 1

# Column widths
for i, w in enumerate(info_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for i in range(num_weeks):
    ws.column_dimensions[get_column_letter(gantt_start_col + i)].width = 3.5

# Row 1: Month headers
hdr_fill = PatternFill(start_color="1F222C", end_color="1F222C", fill_type="solid")
hdr_font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")

for c in range(1, gantt_start_col):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill

current_month = None
month_start = None
for i, wk in enumerate(weeks):
    col = gantt_start_col + i
    mn = wk.strftime("%b %Y")
    cell = ws.cell(row=1, column=col)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center")
    if mn != current_month:
        if month_start is not None and col - 1 > month_start:
            ws.merge_cells(start_row=1, start_column=month_start, end_row=1, end_column=col - 1)
            ws.cell(row=1, column=month_start).value = current_month
        current_month = mn
        month_start = col
if month_start and gantt_start_col + num_weeks - 1 >= month_start:
    ws.merge_cells(start_row=1, start_column=month_start, end_row=1, end_column=gantt_start_col + num_weeks - 1)
    ws.cell(row=1, column=month_start).value = current_month

# Row 2: Week numbers
cream_fill = PatternFill(start_color="F5F3EE", end_color="F5F3EE", fill_type="solid")
for c in range(1, gantt_start_col):
    ws.cell(row=2, column=c).fill = cream_fill

for i, wk in enumerate(weeks):
    col = gantt_start_col + i
    wn = wk.isocalendar()[1]
    cell = ws.cell(row=2, column=col, value=f"W{wn}")
    cell.font = Font(name="Calibri", size=7, color="666666")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

# Row 3: Column headers + date labels
for i, h in enumerate(info_cols, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

for i, wk in enumerate(weeks):
    col = gantt_start_col + i
    cell = ws.cell(row=3, column=col, value=wk.strftime("%d/%m"))
    cell.font = Font(name="Calibri", size=7, color="999999")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin


def week_idx(dt_str):
    if isinstance(dt_str, str):
        dt = datetime.strptime(dt_str, "%Y-%m-%d")
    else:
        dt = dt_str
    delta = (dt - start_date).days // 7
    return max(0, min(delta, num_weeks - 1))


# Tasks: (id, track, task, owner, duration, depends, start, end, color)
tasks = [
    ("", "BPM TRACK", "", "", "", "", None, None, None),
    (1, "Pre-PO", "Supplier qualification Jira (Transtek)", "Andre", "2-4 wk", "-", "2026-04-09", "2026-05-07", "prepo"),
    (2, "Pre-PO", "SQA negotiation (Transtek)", "Bianca/Andre", "2-4 wk", "Parallel", "2026-04-09", "2026-05-07", "prepo"),
    (3, "Design", "Label + IFU + box artwork (BPM)", "Marta/Paulo", "1-2 wk", "-", "2026-04-09", "2026-04-23", "design"),
    (4, "Design", "Confirm cuff branding + BLE name", "Andre", "1 wk", "-", "2026-04-09", "2026-04-16", "design"),
    (5, "Approval", "Golden sample request (Transtek)", "Andre", "1 wk", "#3", "2026-04-23", "2026-04-30", "approval"),
    (6, "Approval", "Golden sample review + sign-off", "Paulo/Andre", "1-2 wk", "#5", "2026-04-30", "2026-05-14", "approval"),
    (7, "Production", "BPM production (12-16 wk)", "Transtek", "12-16 wk", "#1,#2,#6", "2026-05-14", "2026-09-03", "prod_bpm"),
    (8, "FDA", "FDA registration + UDI (BPM)", "Bianca", "~1 wk", "#7", "2026-09-03", "2026-09-10", "fda"),
    (9, "Shipping", "BPM air shipment to SLC", "Andre/Fernando", "7-10 d", "#8", "2026-09-10", "2026-09-20", "prod_bpm"),
    (10, "Shipping", "BPM sea shipment (remaining)", "Andre/Fernando", "6-8 wk", "#8", "2026-09-10", "2026-11-01", "fda"),

    ("", "SCALE TRACK", "", "", "", "", None, None, None),
    (11, "Pre-PO", "Supplier qualification Jira (Unique Scales)", "Andre", "2-4 wk", "-", "2026-04-09", "2026-05-07", "prepo"),
    (12, "Pre-PO", "SQA negotiation (Unique Scales)", "Bianca/Andre", "2-4 wk", "Parallel", "2026-04-09", "2026-05-07", "prepo"),
    (13, "Pre-PO", "Confirm hybrid dual-freq config", "Andre/Queenie", "TBC", "-", "2026-04-09", "2026-04-23", "prepo"),
    (14, "Design", "Label + IFU + box artwork (Scale)", "Marta/Paulo", "1-2 wk", "-", "2026-04-09", "2026-04-23", "design"),
    (15, "Approval", "Golden sample request (Unique Scales)", "Andre", "1 wk", "#14", "2026-04-23", "2026-04-30", "approval"),
    (16, "Approval", "Golden sample review + sign-off", "Paulo/Andre", "1-2 wk", "#15", "2026-04-30", "2026-05-14", "approval"),
    (17, "Production", "Scale production (35-40 days)", "Unique Scales", "35-40 d", "#11-#13,#16", "2026-05-14", "2026-06-23", "prod_scale"),
    (18, "FDA", "FDA registration + UDI (Scale)", "Bianca", "~1 wk", "#17", "2026-06-23", "2026-06-30", "fda"),
    (19, "Shipping", "Scale air shipment to SLC", "Andre/Fernando", "5-7 d", "#18", "2026-06-30", "2026-07-07", "prod_scale"),

    ("", "CROSS-CUTTING", "", "", "", "", None, None, None),
    (20, "Finance", "PLD accounting memo", "Anand", "1 wk", "-", "2026-04-09", "2026-04-16", "finance"),
    (21, "Finance", "Finance team adjustment", "Paulo/Finance", "TBD", "#20", "2026-04-16", "2026-04-30", "finance"),
    (22, "Future", "Glucometer/CGM research", "Andre/Anand", "TBD", "Anand NPI", "2026-05-01", "2026-10-01", "fda"),
]

# Render
row = 4
for t in tasks:
    tid, track, task, owner, dur, dep, sd, ed, color = t

    # Phase separator
    if tid == "" and track:
        for c in range(1, gantt_start_col + num_weeks):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
        ws.cell(row=row, column=2, value=track).font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        row += 1
        continue

    # Task info
    vals = [tid, track, task, owner, dur, dep]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(name="Calibri", size=9)
        cell.border = border_thin
        if c == 3:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(vertical="center")

    # Gantt bars
    for wi in range(num_weeks):
        ws.cell(row=row, column=gantt_start_col + wi).border = border_thin

    if sd and ed:
        si = week_idx(sd)
        ei = week_idx(ed)
        bar_color = colors.get(color, "CCCCCC")
        for wi in range(si, ei + 1):
            col = gantt_start_col + wi
            ws.cell(row=row, column=col).fill = PatternFill(start_color=bar_color, end_color=bar_color, fill_type="solid")

    row += 1

# Today marker (Apr 8 = W15)
today_wi = week_idx(datetime(2026, 4, 8))
today_col = gantt_start_col + today_wi
for r in range(3, row):
    cell = ws.cell(row=r, column=today_col)
    if cell.fill.start_color.index in ["00000000", None, 0]:
        cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    cell.border = Border(
        left=Side(style="medium", color="FF0000"),
        right=cell.border.right,
        top=cell.border.top,
        bottom=cell.border.bottom,
    )

# Legend
row += 1
ws.cell(row=row, column=2, value="LEGEND").font = Font(name="Calibri", bold=True, size=9)
row += 1
legend = [
    ("Pre-PO / Qualification", "prepo"),
    ("Design / Artwork", "design"),
    ("Approval / Golden Sample", "approval"),
    ("BPM Production + Shipping", "prod_bpm"),
    ("Scale Production + Shipping", "prod_scale"),
    ("FDA / Regulatory", "fda"),
    ("Finance / Other", "finance"),
]
for label, c in legend:
    ws.cell(row=row, column=1).fill = PatternFill(start_color=colors[c], end_color=colors[c], fill_type="solid")
    ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=9)
    row += 1

# Milestones
row += 1
ws.cell(row=row, column=2, value="KEY MILESTONES").font = Font(name="Calibri", bold=True, size=9)
row += 1
milestones = [
    ("Scale ready for members", "2026-07-07"),
    ("BPM ready for members (air)", "2026-09-20"),
    ("BPM full delivery (sea)", "2026-11-01"),
]
for label, dt in milestones:
    ws.cell(row=row, column=2, value=label).font = Font(name="Calibri", size=9)
    ws.cell(row=row, column=3, value=dt).font = Font(name="Calibri", size=9, color="D35D5D")
    mi = week_idx(dt)
    cell = ws.cell(row=row, column=gantt_start_col + mi, value="\u25C6")
    cell.font = Font(name="Calibri", size=12, color="D35D5D")
    cell.alignment = Alignment(horizontal="center")
    row += 1

ws.freeze_panes = "G4"

dst = "G:/My Drive/Pulse/General/Pulse_PLD_Gantt.xlsx"
wb.save(dst)
print(f"Saved: {dst}")
